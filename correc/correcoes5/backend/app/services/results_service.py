# app/services/results_service.py
# -*- coding: utf-8 -*-
"""
Serviço de resultados do simulado.

Funções principais:
  - get_class_results()     → dict com notas de todos os alunos de uma turma
  - export_results_xlsx()   → arquivo XLSX por série
  - export_report_pdf()     → PDF devolutiva individual por aluno
"""

from __future__ import annotations
import os
import io
from typing import Optional
from datetime import datetime

from sqlalchemy.orm import Session

# ---------------------------------------------------------------------------
# Cálculo de resultados
# ---------------------------------------------------------------------------

def _calc_nota(acertos: int, total: int) -> float:
    if total == 0:
        return 0.0
    return round((10 / total) * acertos, 2)


def get_student_result(db: Session, exam_id: int, student_id: int) -> dict:
    """
    Retorna resultado completo de um aluno em um simulado.
    Inclui nota geral e breakdown por disciplina.
    """
    from app.models.student_answer import StudentAnswer
    from app.models.exam import ExamQuestionLink, Question
    from app.models.school import Student, SchoolClass
    from app.models.base_models import Discipline

    student = db.get(Student, student_id)
    if not student:
        return {}

    links = (
        db.query(ExamQuestionLink)
        .filter(ExamQuestionLink.exam_id == exam_id)
        .order_by(ExamQuestionLink.order_idx)
        .all()
    )
    total = len(links)

    answers = {
        a.question_link_id: a
        for a in db.query(StudentAnswer)
        .filter_by(exam_id=exam_id, student_id=student_id)
        .all()
    }

    # Por disciplina
    disc_map: dict[int, dict] = {}  # discipline_id → {name, total, acertos}
    per_question = []

    for link in links:
        question = db.get(Question, link.question_id)
        disc_id  = question.discipline_id if question else None

        if disc_id and disc_id not in disc_map:
            disc = db.get(Discipline, disc_id)
            disc_map[disc_id] = {
                "discipline_id":   disc_id,
                "discipline_name": disc.name if disc else str(disc_id),
                "total":           0,
                "acertos":         0,
            }

        answer     = answers.get(link.id)
        marked     = answer.marked_label if answer else None
        is_correct = answer.is_correct   if answer else None

        if disc_id:
            disc_map[disc_id]["total"]   += 1
            if is_correct:
                disc_map[disc_id]["acertos"] += 1

        per_question.append({
            "order_idx":     link.order_idx,
            "question_id":   link.question_id,
            "marked_label":  marked,
            "correct_label": link.correct_label,
            "is_correct":    is_correct,
            "discipline_id": disc_id,
        })

    acertos_total = sum(1 for q in per_question if q["is_correct"])

    for d in disc_map.values():
        d["nota"] = _calc_nota(d["acertos"], d["total"])

    return {
        "student_id":   student_id,
        "student_name": student.name,
        "student_ra":   getattr(student, "ra", "") or getattr(student, "number", "") or "",
        "class_id":     student.class_id,
        "total":        total,
        "acertos":      acertos_total,
        "nota":         _calc_nota(acertos_total, total),
        "por_disciplina": list(disc_map.values()),
        "per_question": per_question,
    }


def get_class_results(db: Session, exam_id: int, class_id: int) -> dict:
    """Retorna resultados de todos os alunos de uma turma, com ranking."""
    from app.models.school import Student

    students = (
        db.query(Student)
        .filter(Student.class_id == class_id)
        .order_by(Student.name)
        .all()
    )

    results = []
    for s in students:
        r = get_student_result(db, exam_id, s.id)
        if r:
            results.append(r)

    # Ranking por nota decrescente
    results.sort(key=lambda x: x["nota"], reverse=True)
    for i, r in enumerate(results):
        r["ranking"] = i + 1

    return {
        "exam_id":  exam_id,
        "class_id": class_id,
        "total_students": len(results),
        "results": results,
    }


# ---------------------------------------------------------------------------
# Exportação XLSX
# ---------------------------------------------------------------------------

def export_results_xlsx(db: Session, exam_id: int, class_id: int) -> bytes:
    """
    Gera planilha XLSX com resultados da turma.
    Colunas: Ranking | Nome | RA | Nota | <disciplina> ... | Q01 | Q02 | ...
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    from app.models.exam import Exam, ExamQuestionLink, Question
    from app.models.school import SchoolClass
    from app.models.base_models import Discipline

    exam = db.get(Exam, exam_id)
    cls  = db.get(SchoolClass, class_id)

    data = get_class_results(db, exam_id, class_id)
    results = data["results"]
    if not results:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Sem resultados para exportar."
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # Coletar disciplinas e questões
    links = (
        db.query(ExamQuestionLink)
        .filter(ExamQuestionLink.exam_id == exam_id)
        .order_by(ExamQuestionLink.order_idx)
        .all()
    )
    disc_ids  = []
    disc_names = {}
    for link in links:
        q = db.get(Question, link.question_id)
        if q and q.discipline_id not in disc_ids:
            disc_ids.append(q.discipline_id)
            disc = db.get(Discipline, q.discipline_id)
            disc_names[q.discipline_id] = disc.name if disc else str(q.discipline_id)

    # Cores
    HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
    DISC_FILL = PatternFill("solid", fgColor="2E75B6")
    Q_FILL    = PatternFill("solid", fgColor="BDD7EE")
    ROW_ALT   = PatternFill("solid", fgColor="F2F2F2")
    GREEN     = PatternFill("solid", fgColor="C6EFCE")
    RED       = PatternFill("solid", fgColor="FFC7CE")
    WHITE_HDR = Font(color="FFFFFF", bold=True, size=10)
    THIN = Side(style="thin", color="AAAAAA")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = cls.name if cls else f"Turma {class_id}"

    # Título
    title_str = f"{exam.title or 'SIMULADO'} — {cls.name if cls else ''} — Devolutiva de Resultados"
    ws.merge_cells(f"A1:{get_column_letter(5 + len(disc_ids) + len(links))}1")
    title_cell = ws["A1"]
    title_cell.value     = title_str
    title_cell.font      = Font(bold=True, size=12, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22

    ws.append([])  # linha 2 vazia

    # Cabeçalho
    headers = ["Ranking", "Nome", "RA", "Nota Geral"]
    for did in disc_ids:
        headers.append(f"Nota {disc_names[did]}")
    for i, link in enumerate(links):
        headers.append(f"Q{i+1:02d}")

    ws.append(headers)
    hdr_row = 3
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=hdr_row, column=col_idx)
        col_name = headers[col_idx - 1]
        if col_idx <= 4:
            cell.fill = HDR_FILL
            cell.font = WHITE_HDR
        elif col_idx <= 4 + len(disc_ids):
            cell.fill = DISC_FILL
            cell.font = WHITE_HDR
        else:
            cell.fill = Q_FILL
            cell.font = Font(bold=True, size=9)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border    = BORDER
    ws.row_dimensions[hdr_row].height = 30

    # Dados
    for r_idx, r in enumerate(results):
        disc_notas = {d["discipline_id"]: d["nota"] for d in r.get("por_disciplina", [])}
        pq_map     = {q["order_idx"]: q for q in r.get("per_question", [])}

        row_data = [r["ranking"], r["student_name"], r["student_ra"], r["nota"]]
        for did in disc_ids:
            row_data.append(disc_notas.get(did, ""))
        for link in links:
            pq = pq_map.get(link.order_idx)
            row_data.append(pq["marked_label"] if pq else "")

        ws.append(row_data)
        data_row = hdr_row + 1 + r_idx
        alt_fill = ROW_ALT if r_idx % 2 == 1 else None

        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=data_row, column=col_idx)
            cell.alignment = Alignment(horizontal="center")
            cell.border    = BORDER

            # Colorir células de questão com certo/errado
            if col_idx > 4 + len(disc_ids):
                link_idx = col_idx - 4 - len(disc_ids) - 1
                if 0 <= link_idx < len(links):
                    link = links[link_idx]
                    pq   = pq_map.get(link.order_idx)
                    if pq and pq["is_correct"] is True:
                        cell.fill = GREEN
                    elif pq and pq["is_correct"] is False:
                        cell.fill = RED
                    elif alt_fill:
                        cell.fill = alt_fill
            elif alt_fill:
                cell.fill = alt_fill

            # Nota geral em negrito
            if col_idx == 4:
                cell.font = Font(bold=True)

    # Larguras
    ws.column_dimensions["A"].width = 9
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 11
    for i in range(len(disc_ids)):
        ws.column_dimensions[get_column_letter(5 + i)].width = 14
    for i in range(len(links)):
        ws.column_dimensions[get_column_letter(5 + len(disc_ids) + i)].width = 5

    ws.freeze_panes = f"A{hdr_row + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# PDF devolutiva individual
# ---------------------------------------------------------------------------

def export_report_pdf(db: Session, exam_id: int, student_id: int) -> bytes:
    """
    Gera PDF devolutiva individual com:
    - Cabeçalho institucional
    - Nota geral e por disciplina
    - Tabela: Nº | Sua Resposta | Gabarito | ✓/✗
    """
    from reportlab.pdfgen.canvas import Canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors

    from app.models.exam import Exam
    from app.models.school import SchoolClass
    from app.services.pdf_generator import (
        STORAGE_BASE, MARGIN_INTERIOR as ML, MARGIN_EXTERIOR as MR,
        _BASE_FONT, _BASE_FONT_BOLD,
    )

    exam    = db.get(Exam, exam_id)
    result  = get_student_result(db, exam_id, student_id)
    if not result:
        raise ValueError("Aluno ou resultado não encontrado.")

    W, H = A4
    buf  = io.BytesIO()
    c    = Canvas(buf, pagesize=A4)

    USABLE = W - ML - MR

    # ── Cabeçalho institucional ────────────────────────────────────
    hdr_top = H - 1.2 * cm
    hdr_bot = H - 4.0 * cm
    c.setLineWidth(0.5)
    c.line(ML, hdr_bot, W - MR, hdr_bot)
    c.line(ML, hdr_top, W - MR, hdr_top)

    logo_w = 2.0 * cm
    logo_h = hdr_top - hdr_bot - 0.4 * cm
    _sp = os.path.join(STORAGE_BASE, "assets", "logo_sp.png")
    if os.path.exists(_sp):
        c.drawImage(_sp, ML, hdr_bot + 0.2*cm, width=logo_w, height=logo_h,
                    preserveAspectRatio=True, mask="auto")
    tx = ML + logo_w + 0.3 * cm
    c.setFont("Times-Bold", 7.5)
    c.drawString(tx, hdr_bot + 2.3*cm, "GOVERNO DO ESTADO DE SÃO PAULO – SECRETARIA DE ESTADO DA EDUCAÇÃO")
    c.setFont(_BASE_FONT, 7.5)
    c.drawString(tx, hdr_bot + 1.8*cm, "UNIDADE REGIONAL DE ENSINO – REGIÃO BAURU – EE PROF. CHRISTINO CABRAL")

    _int = os.path.join(STORAGE_BASE, "assets", "logo_integral.png")
    if os.path.exists(_int):
        c.drawImage(_int, W - MR - 2.2*cm, hdr_bot + 0.2*cm, width=2.0*cm, height=logo_h,
                    preserveAspectRatio=True, mask="auto")

    # ── Título ─────────────────────────────────────────────────────
    y = hdr_bot - 0.8 * cm
    c.setFont("Times-Bold", 13)
    c.drawCentredString(W/2, y, "DEVOLUTIVA DE RESULTADOS")
    y -= 0.5 * cm
    c.setFont(_BASE_FONT, 9)
    c.drawCentredString(W/2, y, (exam.title or "SIMULADO").upper())
    y -= 0.7 * cm

    # ── Dados do aluno ─────────────────────────────────────────────
    BOX_H = 0.55 * cm
    def _box(label, value, bx, by, bw):
        c.setStrokeColorRGB(0.3, 0.3, 0.3)
        c.setFillColorRGB(1, 1, 1)
        c.setLineWidth(0.5)
        c.roundRect(bx, by - BOX_H, bw, BOX_H, 3, fill=1, stroke=1)
        c.setFillColorRGB(0, 0, 0)
        c.setFont("Times-Bold", 7.5)
        lw = c.stringWidth(label + "  ", "Times-Bold", 7.5)
        c.drawString(bx + 0.15*cm, by - BOX_H*0.45, label)
        c.setFont(_BASE_FONT, 8)
        c.drawString(bx + 0.15*cm + lw, by - BOX_H*0.45, value)

    _box("Nome:", result["student_name"], ML, y, USABLE)
    y -= BOX_H + 0.2*cm
    fw = USABLE / 3 - 0.1*cm
    _box("R.A:",   result["student_ra"],  ML,                   y, fw)
    _box("Série:", str(result["class_id"]), ML + fw + 0.15*cm,  y, fw)
    _box("Data:",  datetime.now().strftime("%d/%m/%Y"), ML + 2*(fw+0.15*cm), y, fw)
    y -= BOX_H + 1.4*cm

    # ── Nota geral ─────────────────────────────────────────────────
    nota = result["nota"]
    cor  = (0.0, 0.5, 0.0) if nota >= 5 else (0.8, 0.0, 0.0)

    c.setFillColorRGB(*cor)
    c.setFont("Times-Bold", 28)
    c.drawCentredString(W/2, y, f"{nota:.1f}")
    c.setFont(_BASE_FONT, 8.5)
    c.setFillColorRGB(0, 0, 0)
    c.drawCentredString(W/2, y - 0.5*cm,
                        f"{result['acertos']} acertos de {result['total']} questões")
    y -= 1.5 * cm

    # ── Por disciplina ─────────────────────────────────────────────
    if result["por_disciplina"]:
        c.setFont("Times-Bold", 9)
        c.drawString(ML, y, "Resultado por disciplina:")
        y -= 0.4 * cm
        disc_col_w = USABLE / max(len(result["por_disciplina"]), 1)
        for i, d in enumerate(result["por_disciplina"]):
            dx    = ML + i * disc_col_w
            dnota = d["nota"]
            dcor  = (0.0, 0.5, 0.0) if dnota >= 5 else (0.8, 0.0, 0.0)
            c.setFillColorRGB(0.93, 0.93, 0.93)
            c.roundRect(dx, y - 1.2*cm, disc_col_w - 0.2*cm, 1.2*cm, 4, fill=1, stroke=0)
            c.setFillColorRGB(*dcor)
            c.setFont("Times-Bold", 14)
            c.drawCentredString(dx + disc_col_w/2 - 0.1*cm, y - 0.6*cm, f"{dnota:.1f}")
            c.setFillColorRGB(0.3, 0.3, 0.3)
            c.setFont(_BASE_FONT, 7)
            c.drawCentredString(dx + disc_col_w/2 - 0.1*cm, y - 1.05*cm, d["discipline_name"])
        y -= 1.6 * cm

    # ── Tabela de questões ─────────────────────────────────────────
    c.setLineWidth(0.5)
    c.line(ML, y, W - MR, y)
    y -= 0.4 * cm

    # Cabeçalho da tabela
    COL_Q  = 0.8 * cm
    COL_M  = 2.2 * cm
    COL_G  = 2.2 * cm
    COL_R  = 1.2 * cm
    COL_D  = USABLE - COL_Q - COL_M - COL_G - COL_R
    ROW_H  = 0.52 * cm

    def _tbl_row(qn, marked, correct, ok, disc_name, row_y, is_header=False):
        bg = (0.15, 0.35, 0.6) if is_header else ((1,1,1) if ok is None else ((0.85,0.95,0.85) if ok else (0.98,0.85,0.85)))
        c.setFillColorRGB(*bg)
        c.rect(ML, row_y - ROW_H, USABLE, ROW_H, fill=1, stroke=0)
        # linhas divisórias
        c.setStrokeColorRGB(0.75, 0.75, 0.75)
        c.line(ML, row_y - ROW_H, W - MR, row_y - ROW_H)

        txt_color = (1,1,1) if is_header else (0,0,0)
        c.setFillColorRGB(*txt_color)
        fnt = _BASE_FONT_BOLD if is_header else _BASE_FONT
        sz  = 8
        c.setFont(fnt, sz)

        mid_y = row_y - ROW_H * 0.62
        c.drawCentredString(ML + COL_Q/2,                        mid_y, str(qn))
        c.drawCentredString(ML + COL_Q + COL_M/2,                mid_y, str(marked or "—"))
        c.drawCentredString(ML + COL_Q + COL_M + COL_G/2,        mid_y, str(correct or "—"))
        if not is_header:
            symbol = "✓" if ok else ("✗" if ok is False else "—")
            c.setFillColorRGB(0,0.5,0) if ok else c.setFillColorRGB(0.8,0,0)
            c.setFont(_BASE_FONT_BOLD, 10)
            c.drawCentredString(ML + COL_Q + COL_M + COL_G + COL_R/2, mid_y, symbol)
            c.setFillColorRGB(0.3, 0.3, 0.3)
            c.setFont(_BASE_FONT, 7)
            c.drawString(ML + COL_Q + COL_M + COL_G + COL_R + 0.1*cm, mid_y, disc_name or "")
        else:
            c.setFillColorRGB(1,1,1)
            c.drawCentredString(ML + COL_Q + COL_M + COL_G + COL_R/2, mid_y, str(ok))
            c.drawString(ML + COL_Q + COL_M + COL_G + COL_R + 0.1*cm, mid_y, disc_name or "")

    _tbl_row("Nº", "Sua Resp.", "Gabarito", "Res.", "Disciplina", y, is_header=True)
    y -= ROW_H

    from app.models.exam import Question
    from app.models.base_models import Discipline

    for pq in result["per_question"]:
        if y < 2 * cm:
            c.showPage()
            y = H - 2 * cm

        disc_name = ""
        if pq.get("discipline_id"):
            disc = db.get(Discipline, pq["discipline_id"])
            disc_name = disc.name if disc else ""

        _tbl_row(
            qn        = pq["order_idx"] + 1,
            marked    = pq["marked_label"],
            correct   = pq["correct_label"],
            ok        = pq["is_correct"],
            disc_name = disc_name,
            row_y     = y,
        )
        y -= ROW_H

    # Rodapé
    c.setFont(_BASE_FONT, 7)
    c.setFillColorRGB(0.5, 0.5, 0.5)
    c.drawCentredString(W/2, 1.0*cm, f"SAMBA — Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}")

    c.save()
    return buf.getvalue()


# ---------------------------------------------------------------------------
# ZIP com devolutivas de toda a turma
# ---------------------------------------------------------------------------

def export_class_reports_zip(db: Session, exam_id: int, class_id: int) -> bytes:
    """
    Gera um ZIP com um PDF de devolutiva por aluno da turma.
    Arquivo nomeado: {RA}_devolutiva.pdf
    """
    import zipfile
    from app.models.school import Student

    students = (
        db.query(Student)
        .filter(Student.class_id == class_id)
        .order_by(Student.name)
        .all()
    )

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for student in students:
            ra = str(getattr(student, "ra", "") or getattr(student, "number", "") or str(student.id))
            try:
                pdf_bytes = export_report_pdf(db, exam_id, student.id)
                zf.write.__doc__  # dummy
                zf.writestr(f"{ra}_devolutiva.pdf", pdf_bytes)
            except Exception as e:
                import logging
                logging.getLogger(__name__).warning(
                    f"[results] devolutiva falhou student={student.id}: {e}"
                )

    return buf.getvalue()
